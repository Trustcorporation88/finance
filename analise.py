# -*- coding: utf-8 -*-
"""Módulo de análise financeira - CFO de Bolso.

Pega dados crus (CSV/Excel/texto), categoriza e devolve os números do
Raio-X do caixa. Todo cálculo é feito aqui (sem "chute"); a IA só
interpreta os números prontos.
"""
from __future__ import annotations

import re
import io
from datetime import datetime
from collections import OrderedDict

import pandas as pd


# --------------------------------------------------------------------------
# 1. PARSING: converte arquivo/texto em DataFrame normalizado
# --------------------------------------------------------------------------

VALOR_COLS = ["valor", "value", "valor da conta", "vlr", "amount", "valor da nota", "valor doc", "valor da nf", "valor liquido", "valor a pagar", "valor a receber", "valor do documento", "bruto"]
DATA_COLS = ["data", "date", "lançamento", "lancamento", "dt", "data do documento", "data de pagamento", "vencimento", "dt lancto", "dt. comp"]
DESC_COLS = ["descrição", "descricao", "description", "histórico", "historico", "fornecedor", "nome", "conta", "observação", "observacao", "detalhe", "categoria"]
TIPO_COLS = ["tipo", "type", "natureza", "tipo de lançamento", "tipo do lançamento", "entrada/saída", "debito/credito", "crédito", "credito"]


def _normalizar(nome: str) -> str:
    """Remove acentos e normaliza para minúsculas."""
    if not isinstance(nome, str):
        return ""
    acentos = {
        "á": "a", "à": "a", "â": "a", "ã": "a", "ä": "a",
        "é": "e", "è": "e", "ê": "e", "ë": "e",
        "í": "i", "ì": "i", "î": "i", "ï": "i",
        "ó": "o", "ò": "o", "ô": "o", "õ": "o", "ö": "o",
        "ú": "u", "ù": "u", "û": "u", "ü": "u",
        "ç": "c", "ñ": "n",
    }
    n = nome.lower().strip()
    for a, b in acentos.items():
        n = n.replace(a, b)
    return n


def _encontrar_coluna(colunas, candidatas):
    norm = {_normalizar(c): c for c in colunas if isinstance(c, str)}
    for cand in candidatas:
        nc = _normalizar(cand)
        if nc in norm:
            return norm[nc]
    # busca parcial
    for c in colunas:
        nc = _normalizar(str(c))
        for cand in candidatas:
            if cand in nc:
                return c
    return None


def _parse_valor(v):
    """Converte 'R$ 1.234,56', '-1.234,56', '1,234.56' etc em float."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    s = s.replace("R$", "").replace("$", "").replace(" ", "").replace("\u00a0", "")
    negativo = False
    if s.startswith("(") and s.endswith(")"):
        negativo = True
        s = s[1:-1]
    if s.startswith("-"):
        negativo = True
        s = s[1:]
    # remove separador de milhar que não seja decimal
    if "," in s and "." in s:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        # pt-BR: 1.234,56 -> 1234.56 ; também 1,5 -> 1.5
        if re.fullmatch(r"-?\d{1,3}(\.\d{3})*(,\d+)?", s):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", ".")
    try:
        v = float(s)
    except ValueError:
        return None
    return -v if negativo else v


def _parse_data(v):
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, pd.Timestamp):
        return v.to_pydatetime()
    if isinstance(v, (int, float)):
        try:
            return pd.Timestamp(v).to_pydatetime()
        except Exception:
            return None
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    try:
        return pd.to_datetime(s, errors="coerce").to_pydatetime()
    except Exception:
        return None


def _detectar_tipo_sinal(valor):
    """Positivo = entrada (dinheiro que entrou); negativo = saída."""
    if valor is None:
        return None
    if valor >= 0:
        return "entrada"
    return "saida"


def ler_dataframe(arquivo: bytes, nome_arquivo: str = "") -> pd.DataFrame:
    """Lê um arquivo (CSV/XLSX) e devolve DataFrame normalizado com colunas:
    data, descricao, valor, tipo, categoria."""
    ext = (nome_arquivo or "").lower().rsplit(".", 1)[-1]
    if ext == "xlsx":
        df = pd.read_excel(io.BytesIO(arquivo))
    else:
        raw = arquivo.decode("utf-8-sig", errors="replace")
        sep = ";"
        if raw.count(";") < raw.count(",") and "," in raw:
            sep = ","
        try:
            df = pd.read_csv(io.StringIO(raw), sep=sep, encoding="utf-8-sig")
        except Exception:
            df = pd.read_csv(io.StringIO(raw), sep=sep, encoding="latin-1")
    return _normalizar_df(df)


def _normalizar_df(df: pd.DataFrame) -> pd.DataFrame:
    colunas = list(df.columns)
    col_valor = _encontrar_coluna(colunas, VALOR_COLS)
    col_data = _encontrar_coluna(colunas, DATA_COLS)
    col_desc = _encontrar_coluna(colunas, DESC_COLS)
    col_tipo = _encontrar_coluna(colunas, TIPO_COLS)

    if col_valor is None and df.shape[1] >= 1:
        # tenta a última coluna numérica
        for c in reversed(colunas):
            vals = df[c].dropna()
            if len(vals) and all(isinstance(x, (int, float)) for x in vals[:5]):
                col_valor = c
                break
    if col_valor is None:
        raise ValueError("Não encontrei uma coluna de valores. Procure uma coluna chamada 'valor', 'vlr', 'amount' etc.")

    out = pd.DataFrame()
    out["valor_raw"] = df[col_valor]
    out["valor"] = out["valor_raw"].apply(_parse_valor)
    if col_desc:
        out["descricao"] = df[col_desc].fillna("")
    else:
        out["descricao"] = ""
    out["data"] = (df[col_data] if col_data else pd.Series([None] * len(out), index=out.index))
    if col_data:
        out["data"] = out["data"].apply(_parse_data)
    if col_tipo:
        out["tipo_col"] = df[col_tipo]
        out["tipo_col"] = out["tipo_col"].apply(lambda x: _normalizar(str(x)))
    else:
        out["tipo_col"] = pd.Series([None] * len(out), index=out.index)
    if col_desc:
        out["categoria"] = df[col_desc].fillna("")
    else:
        out["categoria"] = ""
    out = out.dropna(subset=["valor"])
    out = out[out["valor"] != 0]
    out = out.reset_index(drop=True)

    # definição de tipo: usa sinal, mas respeita coluna explícita se existir
    sinais = out["valor"].apply(_detectar_tipo_sinal)
    if col_tipo:
        def tipo_final(row):
            t = str(row["tipo_col"]).lower()
            if any(k in t for k in ["entrada", "receita", "credito", "recebimento", "venda", "+"]):
                return "entrada"
            if any(k in t for k in ["saida", "despesa", "debito", "pagamento", "custo", "-"]):
                return "saida"
            return row["sinal"]
        out["sinal"] = sinais
        out["tipo"] = out.apply(tipo_final, axis=1)
    else:
        out["tipo"] = sinais
    out["categoria"] = out["descricao"]
    return out


def ler_texto(texto: str) -> pd.DataFrame:
    """Interpreta texto colado. Aceita linhas 'data;descrição;valor' ou
    'descrição;valor'. Tenta também interpretar como CSV inline."""
    texto = texto.strip()
    if not texto:
        raise ValueError("Cole algum dado primeiro.")
    if "\n" in texto or (";" in texto and len(texto.split(";")) > 1):
        linhas = [l for l in texto.splitlines() if l.strip()]
        if linhas:
            primeira = linhas[0]
            primeira_min = _normalizar(primeira)
            # se a primeira linha não parece cabeçalho (não tem palavras de coluna),
            # adiciona um cabeçalho padrão
            parece_cabecalho = any(
                kw in primeira_min for kw in ("data", "descri", "valor", "vlr", "date", "description", "value")
            )
            if not parece_cabecalho:
                ncol = len(primeira.split(";"))
                if ncol == 3:
                    texto = "data;descricao;valor\n" + texto
                elif ncol == 2:
                    texto = "descricao;valor\n" + texto
        try:
            return ler_dataframe(texto.encode("utf-8"), "dados.csv")
        except Exception:
            pass
    # formato livre: cada linha "valor, descrição" ou "descrição - valor" ou "valor descrição"
    linhas = []
    for linha in texto.splitlines():
        linha = linha.strip()
        if not linha:
            continue
        # "valor" sozinho
        v = _parse_valor(linha)
        if v is not None:
            linhas.append({"valor": v, "descricao": "", "data": None})
            continue
        # tenta padrões com número dentro
        nums = re.findall(r"-?R?\$?\s?[\d.\.,]{2,}", linha)
        if nums:
            v = _parse_valor(nums[-1])
            desc = linha.replace(nums[-1], "").strip(" -:;,")
            if v is not None:
                linhas.append({"valor": v, "descricao": desc, "data": None})
    if not linhas:
        raise ValueError("Não consegui interpretar o texto. Use formato 'valor;descrição' ou uma planilha.")
    df = pd.DataFrame(linhas)
    return _normalizar_df(df)


# --------------------------------------------------------------------------
# 2. CATEGORIZAÇÃO AUTOMÁTICA
# --------------------------------------------------------------------------

REGRA_CATEGORIA = [
    (r"(freelance|freela|equipe|time|colaborador|salario|salário|folha|pessoa|clt|pj|pro.labore|prolabore|encargo|fgts|inss|vale.alimenta|vale.refe|vr\b|va\b)", "Pessoal"),
    (r"(aluguel|aluguer|locacao|locação|galpão|galpao|condominio|condomínio|iptu|arrenda)", "Locações / Imóvel"),
    (r"(internet|software|licença|licenca|sistema|erp|nuvem|assinatura|ferramenta|ti\b|hosting|domínio|dominio|app\b)", "Tecnologia / Software"),
    (r"(energia|eletricidade|eletrica|agua|água|gás|gas\b|combustivel|combustível)", "Água / Energia / Gás"),
    (r"(transporte|frete|carreto|combustivel|combustível|pedagio|pedágio|veiculo|veículo|uber|taxi|táxi|gasolina|diesel)", "Transporte"),
    (r"(imposto|tributo|taxa|iss\b|pis\b|cofins|icms|irpj|csll|dar[íf]|nota fiscal|guia)", "Impostos"),
    (r"(comiss[ãa]o|comissao|bônus|bonus|premia|premiacao|premiação|incentivo)", "Comissões / Bônus"),
    (r"(marketing|anuncio|anúncio|traffego|tráfego|impulsionamento|ads\b|google ads|meta ads|publicidade)", "Marketing / Tráfego"),
    (r"(contador|contabilidade|advogad|juridic|consultor|consultoria|assessoria|serviço de terceiro|servico de terceiro)", "Serviços de Terceiros"),
    (r"(manutencao|manutenção|reparo|conserto|instalac|instalação|predial|obra)", "Manutenção / Obras"),
    (r"(viagem|passagem|hotel|diaria|diária|deslocamento|hospedagem|aéreo|aereo|vo[oô])", "Viagens"),
    (r"(cartao|cartão|juros|multa|taxa banc|tarifa|anuuidade|anuidade|cheque especial)", "Financeiro / Banco"),
    (r"(mercado|supermercado|mercado|alimentaç|alimentac|supermercado|compras|material de escritorio)", "Compras / Suprimentos"),
    (r"(cliente|devolução|devolucao|reembolso|ressarcimento|estorno|chargeback)", "Devoluções / Reembolsos"),
]

CATEGORIA_OUTROS = "Outros"


def categorizar(df: pd.DataFrame) -> pd.DataFrame:
    """Preenche a coluna categoria a partir da descrição."""
    if "descricao" not in df.columns:
        df["descricao"] = ""
    def cat(desc):
        d = _normalizar(desc)
        for padrao, nome in REGRA_CATEGORIA:
            if re.search(padrao, d):
                return nome
        return CATEGORIA_OUTROS
    df["categoria"] = df["descricao"].apply(cat)
    return df


# --------------------------------------------------------------------------
# 3. CÁLCULOS DO RAIO-X
# --------------------------------------------------------------------------

def calcular(df: pd.DataFrame) -> dict:
    """Calcula todos os números do Raio-X a partir do DataFrame."""
    df = df.copy()
    df["categoria"] = df["categoria"].fillna(CATEGORIA_OUTROS)
    df = categorizar(df) if "categoria" not in df.columns or df["categoria"].isna().all() else df

    entradas = df[df["tipo"] == "entrada"]
    saidas = df[df["tipo"] == "saida"]

    total_entradas = entradas["valor"].sum()
    total_saidas = abs(saidas["valor"].sum())
    saldo = total_entradas - total_saidas
    margem = (saldo / total_entradas * 100) if total_entradas else None

    gastos_cat = saidas.groupby("categoria")["valor"].sum().abs().sort_values(ascending=False)
    entradas_cat = entradas.groupby("categoria")["valor"].sum().abs().sort_values(ascending=False)

    # por mês
    if "data" in df.columns and df["data"].notna().any():
        df_mes = df.copy()
        df_mes["mes"] = df_mes["data"].dt.to_period("M").astype(str)
        por_mes = {}
        for mes, grupo in df_mes.groupby("mes"):
            por_mes[mes] = {
                "entradas": round(grupo[grupo["tipo"] == "entrada"]["valor"].sum(), 2),
                "saidas": round(abs(grupo[grupo["tipo"] == "saida"]["valor"].sum()), 2),
                "saldo": round(grupo[grupo["tipo"] == "entrada"]["valor"].sum()
                               - abs(grupo[grupo["tipo"] == "saida"]["valor"].sum()), 2),
            }
    else:
        por_mes = None

    n_entradas = len(entradas)
    n_saidas = len(saidas)

    # alertas automáticos
    alertas = []
    if saldo < 0:
        alertas.append({
            "nivel": "alto",
            "titulo": "Caixa negativo no período",
            "detalhe": f"Saíram R$ {total_saidas:,.0f} e entraram R$ {total_entradas:,.0f} — faltaram R$ {abs(saldo):,.0f}. Está gastando mais do que entra.",
        })
    elif total_entradas == 0:
        alertas.append({
            "nivel": "alto",
            "titulo": "Sem entradas identificadas",
            "detalhe": "Só encontrei saídas nos dados. Confira se as entradas (recebimentos/vendas) estão no arquivo — não vou chutar receita.",
        })
    if margem is not None and margem < 10:
        alertas.append({
            "nivel": "medio",
            "titulo": "Margem apertada",
            "detalhe": f"Margem de {margem:.1f}%. Sobra pouco de cada R$ 100 — fôlego baixo para imprevistos.",
        })
    if len(gastos_cat) and gastos_cat.iloc[0] / total_saidas > 0.5:
        alertas.append({
            "nivel": "medio",
            "titulo": "Concentração de gasto",
            "detalhe": f"'{gastos_cat.index[0]}' responde por {gastos_cat.iloc[0] / total_saidas * 100:.0f}% das saídas — confira se não há gordura aí.",
        })
    # concentração de receita
    if len(entradas_cat) and total_entradas > 0 and entradas_cat.iloc[0] / total_entradas > 0.3:
        alertas.append({
            "nivel": "alto",
            "titulo": "Concentração de receita",
            "detalhe": f"'{entradas_cat.index[0]}' representa {entradas_cat.iloc[0] / total_entradas * 100:.0f}% das entradas. Se essa fonte falhar, o caixa sofre.",
        })

    # validações de consistência (divergências potenciais)
    divergencias = []
    # 1. Soma das categorias de saída vs total de saídas
    soma_categorias = gastos_cat.sum()
    if total_saidas > 0 and abs(soma_categorias - total_saidas) > max(1.0, total_saidas * 0.001):
        divergencias.append({
            "nivel": "alto",
            "titulo": "Categorias não fecham com o total de saídas",
            "detalhe": f"As categorias somam R$ {soma_categorias:,.2f}, mas o total de saídas é R$ {total_saidas:,.2f} (diferença de R$ {total_saidas - soma_categorias:,.2f}). Confira se há lançamentos sem categoria.",
        })
    # 2. Valores sem tipo definido
    sem_tipo = df[df["tipo"].isna()]
    if len(sem_tipo):
        divergencias.append({
            "nivel": "medio",
            "titulo": "Lançamentos sem classificação entrada/saída",
            "detalhe": f"{len(sem_tipo)} lançamento(s) não foram classificados (valor zero ou sem sinal). Eles não entraram no cálculo.",
        })
    # 3. Checagem de saldo mensal (saldo anterior + entradas - saídas = saldo final)
    if por_mes:
        meses_ordenados = sorted(por_mes.keys())
        if len(meses_ordenados) >= 2:
            for i in range(1, len(meses_ordenados)):
                mes_ant = meses_ordenados[i - 1]
                mes_atual = meses_ordenados[i]
                saldo_ant = por_mes[mes_ant]["saldo"]
                saldo_esperado = saldo_ant + por_mes[mes_atual]["entradas"] - por_mes[mes_atual]["saidas"]
                saldo_real = por_mes[mes_atual]["saldo"]
                if abs(saldo_esperado - saldo_real) > max(1.0, abs(saldo_esperado) * 0.001):
                    divergencias.append({
                        "nivel": "alto",
                        "titulo": f"Saldo de {mes_atual} não fecha com o mês anterior",
                        "detalhe": f"{mes_ant} fechou em R$ {saldo_ant:,.2f}; com as entradas/saídas de {mes_atual}, o saldo esperado é R$ {saldo_esperado:,.2f}, mas o registrado é R$ {saldo_real:,.2f} (diferença de R$ {saldo_real - saldo_esperado:,.2f}). Pode haver receita faltando ou lançamento errado.",
                    })
    # 4. Categoria 'Outros' muito grande
    if "Outros" in gastos_cat and total_saidas > 0:
        pct_outros = gastos_cat["Outros"] / total_saidas * 100
        if pct_outros > 40:
            divergencias.append({
                "nivel": "medio",
                "titulo": "Muitos gastos sem categoria ('Outros')",
                "detalhe": f"'Outros' representa {pct_outros:.0f}% das saídas. Renomeie as descrições para categorizar melhor.",
            })
    # 5. Lançamentos com valor muito alto em relação à mediana (possível erro de digitação)
    if len(saidas) >= 4 and total_saidas > 0:
        mediana = saidas["valor"].abs().median()
        if mediana > 0:
            outliers = saidas[saidas["valor"].abs() > mediana * 5]
            if len(outliers):
                maiores = outliers.sort_values("valor").head(3)
                for _, row in maiores.iterrows():
                    desc = str(row.get("descricao", ""))[:60] or "(sem descrição)"
                    divergencias.append({
                        "nivel": "medio",
                        "titulo": "Possível erro de digitação ou valor atípico",
                        "detalhe": f"'{desc}' = R$ {abs(row['valor']):,.2f} — muito acima da mediana de R$ {mediana:,.2f}. Confira.",
                    })
    # 6. Lançamentos duplicados (mesma data + valor + descrição)
    try:
        df_dup = df.copy()
        chave = ["valor"]
        if "data" in df.columns and df["data"].notna().any():
            chave.append("data")
        if "descricao" in df.columns:
            chave.append("descricao")
        duplicados = df_dup[df_dup.duplicated(subset=chave, keep=False)]
        if len(duplicados):
            pares = duplicados.groupby(chave).size().sort_values(ascending=False)
            exemplos = []
            for idx, contagem in list(pares.items())[:3]:
                v = idx[0] if isinstance(idx, tuple) else duplicados.iloc[0]["valor"]
                desc = duplicados.iloc[0].get("descricao", "") if len(chave) == 1 else idx[-1]
                exemplos.append(f"'{str(desc)[:40]}' = R$ {abs(v):,.2f} ({contagem}x)")
            divergencias.append({
                "nivel": "medio",
                "titulo": "Lançamentos duplicados no extrato",
                "detalhe": f"{len(duplicados)} linha(s) parecem duplicadas (mesma data+valor+descrição). Ex.: " + "; ".join(exemplos) + ". Confira se o extrato não foi importado 2x.",
            })
    except Exception:
        pass

    # 7. Detecção de unidade (valores misturados: R$, milhares, milhões)
    if total_entradas > 0 or total_saidas > 0:
        valores_abs = df["valor"].abs()
        qtd_baixa = int((valores_abs < 1000).sum())
        qtd_mil = int(((valores_abs >= 1000) & (valores_abs < 1000000)).sum())
        qtd_milhao = int((valores_abs >= 1000000).sum())
        if qtd_milhao > 0 and qtd_baixa > 0 and len(df) >= 6:
            pct_baixo = qtd_baixa / len(df) * 100
            if pct_baixo > 10:
                divergencias.append({
                    "nivel": "alto",
                    "titulo": "Valores possivelmente em unidades misturadas",
                    "detalhe": f"Encontrei {qtd_milhao} lançamento(s) de R$ 1 milhão+ junto com {qtd_baixa} lançamentos pequenos. Confira se alguns valores não estão em 'mil' (ex.: 1.500 = R$ 1,5 mil ou R$ 1.500?) — isso distorce o total.",
                })

    # 8. Resumo de conferência de dados
    sem_data = int(df["data"].isna().sum()) if "data" in df.columns else len(df)
    sem_desc = int(df["descricao"].fillna("").astype(str).str.strip().eq("").sum()) if "descricao" in df.columns else 0

    conferencia = {
        "lidos": int(len(df)),
        "entradas": n_entradas,
        "saidas": n_saidas,
        "sem_data": sem_data,
        "sem_descricao": sem_desc,
        "observacao": "",
    }
    if sem_data == len(df):
        conferencia["observacao"] = "Nenhum lançamento tem data — não foi possível montar o fluxo por mês (só totais)."
    elif sem_data > 0:
        conferencia["observacao"] = f"{sem_data} lançamento(s) sem data não entraram no fluxo mensal."

    return {
        "total_entradas": round(total_entradas, 2),
        "total_saidas": round(total_saidas, 2),
        "saldo": round(saldo, 2),
        "margem": round(margem, 1) if margem is not None else None,
        "gastos_categorias": {str(k): round(v, 2) for k, v in gastos_cat.items()},
        "entradas_categorias": {str(k): round(v, 2) for k, v in entradas_cat.items()},
        "por_mes": por_mes,
        "n_entradas": n_entradas,
        "n_saidas": n_saidas,
        "n_transacoes": len(df),
        "alertas": alertas,
        "divergencias": divergencias,
        "conferencia": conferencia,
        "categorias": {str(k): round(v, 2) for k, v in gastos_cat.items()},
    }


def resumo_para_ia(resultado: dict) -> str:
    """Monta o texto-resumo dos números para a IA interpretar."""
    linhas = [
        "NÚMEROS CALCULADOS (regime de caixa, sem estimativa):",
        f"- Total de entradas: R$ {resultado['total_entradas']:,.2f}",
        f"- Total de saídas: R$ {resultado['total_saidas']:,.2f}",
        f"- Saldo do período: R$ {resultado['saldo']:,.2f}",
    ]
    if resultado["margem"] is not None:
        linhas.append(f"- Margem do período: {resultado['margem']:.1f}%")
    else:
        linhas.append("- Margem: não calculável (sem entradas).")
    linhas.append(f"- Transações: {resultado['n_transacoes']} ({resultado['n_entradas']} entradas, {resultado['n_saidas']} saídas)")
    if resultado["gastos_categorias"]:
        linhas.append("- Saídas por categoria:")
        total_s = resultado["total_saidas"] or 1
        for cat, val in resultado["gastos_categorias"].items():
            linhas.append(f"    * {cat}: R$ {val:,.2f} ({val / total_s * 100:.1f}%)")
    if resultado["entradas_categorias"]:
        linhas.append("- Entradas por categoria:")
        total_e = resultado["total_entradas"] or 1
        for cat, val in resultado["entradas_categorias"].items():
            linhas.append(f"    * {cat}: R$ {val:,.2f} ({val / total_e * 100:.1f}%)")
    if resultado["por_mes"]:
        linhas.append("- Fluxo por mês:")
        for mes, v in resultado["por_mes"].items():
            linhas.append(f"    * {mes}: entradas R$ {v['entradas']:,.2f} | saídas R$ {v['saidas']:,.2f} | saldo R$ {v['saldo']:,.2f}")
    if resultado["alertas"]:
        linhas.append("- Alertas automáticos detectados:")
        for a in resultado["alertas"]:
            linhas.append(f"    * [{a['nivel'].upper()}] {a['titulo']}: {a['detalhe']}")
    if resultado.get("divergencias"):
        linhas.append("- Divergências/validações de consistência:")
        for d in resultado["divergencias"]:
            linhas.append(f"    * [{d['nivel'].upper()}] {d['titulo']}: {d['detalhe']}")
    return "\n".join(linhas)


# --------------------------------------------------------------------------
# 4. LEITURA AVANÇADA: planilhas complexas (modelos com várias abas)
# --------------------------------------------------------------------------

def ler_planilha_completa(arquivo: bytes, nome_arquivo: str = "") -> dict:
    """Lê uma planilha completa (todas as abas) e devolve um resumo estruturado.

    Serve para arquivos que não são extratos simples (ex.: modelos com várias
    abas, orçamentos, DRE, planos financeiros). Devolve um dict com as abas,
    colunas, amostras de dados e totais — pronto para a IA interpretar.
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(arquivo), data_only=True, read_only=True)
    abas_info = []
    for sn in wb.sheetnames:
        ws = wb[sn]
        linhas = []
        max_cols = 0
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 30:
                break
            vals = []
            for v in row:
                if v is None:
                    vals.append("")
                elif isinstance(v, float) and v == int(v):
                    vals.append(str(int(v)))
                else:
                    vals.append(str(v))
            vals = [v for v in vals if v != ""]
            if not vals:
                continue
            linhas.append(vals)
            max_cols = max(max_cols, len(vals))
        # extrai cabeçalhos (primeiras linhas não vazias)
        cabecalhos = []
        for l in linhas[:5]:
            cabecalhos.append(l)
        # amostra de dados (linhas depois dos cabeçalhos)
        amostra = linhas[5:15]
        abas_info.append({
            "nome": sn,
            "linhas_lidas": len(linhas),
            "colunas_max": max_cols,
            "cabecalhos": cabecalhos,
            "amostra": amostra,
        })
    wb.close()

    # monta o texto-resumo da planilha para a IA
    texto = (
        f"ARQUIVO: {nome_arquivo or 'planilha.xlsx'}\n"
        f"É uma PLANILHA COMPLETA com {len(abas_info)} aba(s). Não é um extrato simples "
        "de transações; é um modelo/relatório com múltiplas abas.\n\n"
        "ESTRUTURA DAS ABAS:\n"
    )
    for a in abas_info:
        texto += f"\n--- ABA: {a['nome']} ({a['linhas_lidas']} linhas, {a['colunas_max']} colunas) ---\n"
        texto += "Cabeçalhos/detalhes:\n"
        for c in a["cabecalhos"]:
            texto += "  " + " | ".join(c) + "\n"
        if a["amostra"]:
            texto += "Amostra de dados:\n"
            for c in a["amostra"]:
                texto += "  " + " | ".join(c) + "\n"

    return {
        "tipo": "planilha_completa",
        "n_abas": len(abas_info),
        "abas": abas_info,
        "resumo_texto": texto,
    }



# --------------------------------------------------------------------------
# 5. FUNÇÕES AVANÇADAS: aba específica, comparação, previsão
# --------------------------------------------------------------------------

def ler_aba_especifica(arquivo: bytes, nome_arquivo: str, aba_nome: str) -> dict:
    """Lê apenas uma aba específica de uma planilha."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(arquivo), data_only=True, read_only=True)
    abas = wb.sheetnames
    if aba_nome not in abas:
        wb.close()
        raise ValueError(f"Aba '{aba_nome}' não encontrada. Abas disponíveis: {', '.join(abas[:10])}")
    ws = wb[aba_nome]
    linhas = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 60:
            break
        vals = []
        for v in row:
            if v is None:
                vals.append("")
            elif isinstance(v, float) and v == int(v):
                vals.append(str(int(v)))
            else:
                vals.append(str(v))
        vals = [v for v in vals if v != ""]
        if vals:
            linhas.append(vals)
    wb.close()

    texto = f"ABA: {aba_nome} ({len(linhas)} linhas)\n"
    for l in linhas[:40]:
        texto += "  " + " | ".join(l)[:200] + "\n"

    return {
        "tipo": "aba_especifica",
        "aba": aba_nome,
        "linhas": linhas[:40],
        "resumo_texto": texto,
    }


def comparar_planilhas(arquivo1: bytes, nome1: str, arquivo2: bytes, nome2: str) -> dict:
    """Lê duas planilhas e monta estrutura comparativa para a IA."""
    import openpyxl

    def ler_resumo(arquivo, nome):
        wb = openpyxl.load_workbook(io.BytesIO(arquivo), data_only=True, read_only=True)
        out = [f"ARQUIVO: {nome}"]
        for sn in wb.sheetnames[:15]:
            ws = wb[sn]
            linhas = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i > 20:
                    break
                vals = [str(v) for v in row if v is not None and str(v) != ""]
                if vals:
                    linhas.append(" | ".join(vals)[:150])
            out.append(f"--- ABA {sn} ({len(linhas)} linhas) ---")
            out.extend(linhas[:15])
        wb.close()
        return "\n".join(out)

    r1 = ler_resumo(arquivo1, nome1)
    r2 = ler_resumo(arquivo2, nome2)
    texto = (
        "COMPARAÇÃO DE DUAS PLANILHAS.\n\n"
        "PLANILHA 1:\n" + r1 + "\n\nPLANILHA 2:\n" + r2
    )
    return {"tipo": "comparacao", "resumo_texto": texto, "planilha1": r1, "planilha2": r2}


def preparar_previsao(resultado: dict) -> str:
    """Monta contexto de previsão para os próximos 3 meses a partir do resultado."""
    linhas = []
    if resultado.get("por_mes"):
        linhas.append("Fluxo mensal real disponível:")
        for mes, v in resultado["por_mes"].items():
            linhas.append(f"  {mes}: entradas R$ {v['entradas']:,.2f} | saídas R$ {v['saidas']:,.2f} | saldo R$ {v['saldo']:,.2f}")
    linhas.append(f"Total entradas: R$ {resultado.get('total_entradas',0):,.2f}")
    linhas.append(f"Total saídas: R$ {resultado.get('total_saidas',0):,.2f}")
    linhas.append(f"Saldo: R$ {resultado.get('saldo',0):,.2f}")
    if resultado.get("gastos_categorias"):
        linhas.append("Saídas por categoria:")
        for c, v in list(resultado["gastos_categorias"].items())[:8]:
            linhas.append(f"  {c}: R$ {v:,.2f}")
    return "\n".join(linhas)
